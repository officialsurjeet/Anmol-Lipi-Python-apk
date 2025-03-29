from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.button import Button
from openpyxl import load_workbook

class ExcelReaderApp(App):
    def build(self):
        self.layout = BoxLayout(orientation='vertical')
        
        # File chooser button
        self.file_chooser_button = Button(text='Select Excel File', size_hint=(1, 0.1))
        self.file_chooser_button.bind(on_press=self.open_file_chooser)
        self.layout.add_widget(self.file_chooser_button)
        
        # Placeholder for file chooser
        self.file_chooser = None
        
        return self.layout

    def open_file_chooser(self, instance):
        if self.file_chooser:
            self.layout.remove_widget(self.file_chooser)
        self.file_chooser = FileChooserIconView()
        self.file_chooser.bind(on_selection=self.load_excel_file)
        self.layout.add_widget(self.file_chooser)

    def load_excel_file(self, filechooser, selection):
        selected_file = selection[0] if selection else None
        if selected_file:
            self.layout.remove_widget(self.file_chooser)
            self.file_chooser = None
            self.display_excel_content(selected_file)

    def display_excel_content(self, file_path):
        try:
            # Load the Excel file
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Iterate through each cell in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        for char_index, char in enumerate(cell.value):
                            # Display each character and its font information
                            font_name = cell.font.name
                            label = Label(text=f'Character: {char}, Font: {font_name}')
                            self.layout.add_widget(label)

        except Exception as e:
            # Show an error message if something goes wrong
            self.layout.add_widget(Label(text=f'Error: {str(e)}'))

if __name__ == '__main__':
    ExcelReaderApp().run()
